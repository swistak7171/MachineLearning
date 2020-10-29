from openpyxl import Workbook
import random
import string
import requests
from bs4 import BeautifulSoup
import re

workbook = Workbook()


def get_random_code() -> string:
    return ''.join(random.choices(string.ascii_lowercase, k=3))


def get_url(code: string) -> string:
    return f"https://stooq.pl/q/?s={code}"


# GIEŁDA
htmls = {}
while len(htmls) < 5:
    code = get_random_code()
    url = get_url(code)
    response = requests.get(url)
    if 'Wyszukiwanie symbolu' not in response.text:
        htmls[code] = response.text

gielda_sheet = workbook.active
gielda_sheet.title = 'Giełda'
gielda_sheet['A1'] = 'Kod'
gielda_sheet['B1'] = 'Kurs'
gielda_sheet['C1'] = 'Zmiana'
gielda_sheet['D1'] = 'Transakcje'

index = 1
for code in htmls:
    index = index + 1
    soup = BeautifulSoup(htmls[code], 'html.parser')
    price_tag = soup.find('span', id=re.compile(f"aq_{code}_c\d"))
    print(code)
    change_tag = soup.find('span', id=f"aq_{code}_m3")
    transactions_tag = soup.find('span', id=f"aq_{code}_n1")

    gielda_sheet[f"A{index}"] = code.upper()
    gielda_sheet[f"B{index}"] = price_tag.text
    gielda_sheet[f"C{index}"] = change_tag.text[1:-1]
    gielda_sheet[f"D{index}"] = transactions_tag.text


# LINKI
linki_sheet = workbook.create_sheet('Linki')
response = requests.get('https://www.pap.pl/')
html = response.text
soup = BeautifulSoup(html, 'html.parser')
links = soup.find_all('a')
hrefs = []
for link in links:
    href = link.get('href')
    if href is not None and href != '':
        hrefs.append(href)

index = 1
for href in hrefs:
    linki_sheet[f"A{index}"] = href
    index = index + 1


# Filmweb
filmweb_sheet = workbook.create_sheet('Filmweb')
filmweb_html = requests.get('https://www.filmweb.pl/film/Szeregowiec+Ryan-1998-179').text
soup = BeautifulSoup(filmweb_html, 'html.parser')
director_tag = soup.find('a', attrs={'itemprop': 'director'})
info_divs = soup.find_all('div', attrs={'class': 'filmInfo__info'})
rate_tag = soup.find('span', attrs={'class': 'filmRating__rateValue'})
date = ''
boxoffice = ''
for div in info_divs:
    for child in div.children:
        try:
            if child.text.startswith('$') and boxoffice == '':
                boxoffice = child.text
        except:
            pass

        if child.name == 'a':
            href = child.attrs['href']
            if href.endswith('dates'):
                date = child.text

filmweb_sheet['A1'] = 'Tytuł'
filmweb_sheet['A2'] = 'Szeregowiec Ryan'
filmweb_sheet['B1'] = 'Reżyser'
filmweb_sheet['B2'] = director_tag.text.strip()
filmweb_sheet['C1'] = 'Data premiery'
filmweb_sheet['C2'] = date.strip()
filmweb_sheet['D1'] = 'Boxoffice'
filmweb_sheet['D2'] = boxoffice.strip()
filmweb_sheet['E1'] = 'Ocena'
filmweb_sheet['E2'] = rate_tag.text.strip()

workbook.save('Szustak-175ICB2.xlsx')